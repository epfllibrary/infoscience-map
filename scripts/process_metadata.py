import yaml
import os
from collections import OrderedDict, defaultdict
from openpyxl import load_workbook

input_file_path = "infoscience-map.xlsx"
output_dir = "../_data/elements"
order_file_path = "../_data/order.yml"

os.makedirs(output_dir, exist_ok=True)

class OrderedDumper(yaml.SafeDumper):
    pass

def _dict_representer(dumper, data):
    return dumper.represent_dict(data.items())

OrderedDumper.add_representer(OrderedDict, _dict_representer)

def generate_yaml(data, index):
    indexes_values = [value.strip() for value in (data.get("indexes", "") or "").split(",")]
    range_values = [value.strip() for value in (data.get("range_values", "") or "").split(",")]
    filename = data.get("name", f"output_{index}")

    yaml_content = OrderedDict(
        [
            ("label", data.get("label", "")),
            ("label-fr", data.get("label-fr", "")),
            ("name", data.get("name", "")),
            ("schema", data.get("schema", "")),
            ("dc-element", data.get("dc-element", "")),
            ("dc-qualifier", data.get("dc-qualifier", "")),
            ("definition", data.get("definition", "")),
            ("obligation", data.get("obligation", "")),
            ("type", data.get("type", "")),
            ("repeatable", data.get("repeatable", "")),
            ("legacy_marc", data.get("legacy_marc", "")),
            (
                "indexes",
                [
                    OrderedDict(
                        [
                            ("configuration", data.get("configuration", "")),
                            ("indexes", indexes_values),
                        ]
                    )
                ],
            ),
            (
                "range",
                [
                    OrderedDict(
                        [
                            ("label", data.get("range_label", "")),
                            ("uri", data.get("range_url", "")),
                            ("values", range_values),
                        ]
                    )
                ],
            ),
        ]
    )

    output_file_path = os.path.join(output_dir, f"{filename}.yaml")

    with open(output_file_path, "w", encoding="utf-8") as yaml_file:
        yaml.dump(
            yaml_content,
            yaml_file,
            Dumper=OrderedDumper,
            default_flow_style=False,
            allow_unicode=True,
        )


def generate_order_yaml(rows):
    order_dict = defaultdict(list)
    for row in rows:
        mtype = row.get("type", "undefined")
        name = row.get("name", "")
        order_dict[mtype].append(name)

    ordered_content = OrderedDict()
    for mtype, names in order_dict.items():
        ordered_content[mtype] = names

    with open(order_file_path, "w", encoding="utf-8") as yaml_file:
        yaml.dump(
            ordered_content,
            yaml_file,
            Dumper=OrderedDumper,
            default_flow_style=False,
            allow_unicode=True,
        )

# Load data from the XLSX file
rows = []
wb = load_workbook(filename=input_file_path, read_only=True)
ws = wb.active

# Assuming the first row contains headers
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

# Read data row by row
for row in ws.iter_rows(min_row=2, values_only=True):
    data = {headers[i]: row[i] for i in range(len(headers))}
    rows.append(data)

# Generate YAML files and order YAML
for index, row in enumerate(rows):
    generate_yaml(row, index)

generate_order_yaml(rows)
